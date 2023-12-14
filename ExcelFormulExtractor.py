import openpyxl


workbook = openpyxl.load_workbook('formulation.xlsx')

sheets = workbook.sheetnames


for sheet_name in sheets:
    sheet = workbook[sheet_name]

    # Dosyayı açın veya oluşturun (sayfa adıyla)
    txt_filename = f'{sheet_name}.txt'
    with open(txt_filename, 'w', encoding='utf-8') as file:
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for column_number, cell_value in enumerate(row, start=1):
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    cell_address = openpyxl.utils.get_column_letter(column_number) + str(row_number)
                    file.write(f"{cell_address} = {cell_value}\n")

workbook.close()
